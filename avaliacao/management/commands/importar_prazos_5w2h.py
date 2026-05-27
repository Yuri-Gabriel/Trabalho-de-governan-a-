import re
from datetime import timedelta

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.utils import timezone

from avaliacao.models import Avaliacao, Resposta, RespostaEscolha


def _parse_when_to_date(when_text: str, base_date):
    """Converte o campo textual "When" da planilha em uma data de calendário (data_limite).

    Regras:
    - "Primeiros 30 dias" / "Até 45 dias" / "Até 60 dias" / "Até 90 dias" -> base_date + N
    - "Imediato" -> base_date
    - "Mensalmente" -> base_date + 30
    - "Trimestralmente" -> base_date + 90
    - "Semestralmente" -> base_date + 180
    - "Contínuo" -> base_date + 365 (marcador de 1 ano)
    - "24x7" -> base_date

    Observação: essas regras são aproximações para permitir uma data objetiva no sistema.
    """
    if not when_text:
        return None

    s = str(when_text).strip().lower()

    if s == "imediato":
        return base_date

    # recorrências aproximadas
    if s == "mensalmente":
        return base_date + timedelta(days=30)
    if s == "a cada trimestre" or s == "trimestralmente":
        return base_date + timedelta(days=90)
    if s == "semestralmente":
        return base_date + timedelta(days=180)
    if s == "contínuo" or s == "continuo":
        return base_date + timedelta(days=365)
    if s == "24x7" or s == "24/7":
        return base_date

    m = re.search(r"(\d+)\s*dias", s)
    if m:
        days = int(m.group(1))
        return base_date + timedelta(days=days)

    return None


class Command(BaseCommand):
    help = "Importa a coluna When (texto) de uma planilha 5W2H e converte para data_limite (calendário) nos planos de ação."

    def add_arguments(self, parser):
        parser.add_argument("--avaliacao-id", type=int, required=True)
        parser.add_argument("--xlsx", type=str, required=True)
        parser.add_argument("--dry-run", action="store_true")

    def handle(self, *args, **options):
        avaliacao_id = options["avaliacao_id"]
        xlsx_path = options["xlsx"]
        dry_run = options["dry_run"]

        try:
            av = Avaliacao.objects.get(id=avaliacao_id)
        except Avaliacao.DoesNotExist as e:
            raise CommandError(f"Avaliacao {avaliacao_id} não encontrada") from e

        base_date = av.criada_em.date() if av.criada_em else timezone.localdate()

        # Respostas NAO (base do plano)
        respostas = list(
            Resposta.objects.filter(avaliacao=av, resposta=RespostaEscolha.NAO)
            .exclude(providencia="")
            .select_related("questao", "plano_acao")
        )
        mapa = {r.questao.texto.strip(): r for r in respostas}

        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb[wb.sheetnames[0]]

        updated = 0
        skipped_non_convertible = 0
        missing = 0

        for row in range(3, ws.max_row + 1):
            what = ws.cell(row, 2).value
            when_text = ws.cell(row, 4).value
            if not what:
                continue

            what = str(what).strip()
            r = mapa.get(what)
            if not r:
                missing += 1
                continue

            plano = getattr(r, "plano_acao", None)
            if not plano:
                continue

            target_date = _parse_when_to_date(when_text, base_date)
            if not target_date:
                skipped_non_convertible += 1
                continue

            if dry_run:
                self.stdout.write(f"[DRY] plano_id={plano.id} when='{when_text}' -> {target_date}")
            else:
                plano.data_limite = target_date
                plano.save(update_fields=["data_limite", "atualizado_em"])

            updated += 1

        self.stdout.write(
            self.style.SUCCESS(
                f"OK: updated={updated} skipped_non_convertible={skipped_non_convertible} missing_questions={missing} base_date={base_date}"
            )
        )
