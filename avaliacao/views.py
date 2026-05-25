import io

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.views.decorators.http import require_http_methods, require_POST

from .decorators import role_required
from .forms import (
    AvaliacaoForm,
    CategoriaQuestaoForm,
    EmpresaForm,
    MetaIndicadorForm,
    ObjetivoEstrategicoPDTIForm,
    PDTIForm,
    PlanoAcaoForm,
    PlanoAcaoInlineForm,
    QuestaoForm,
    RespostaForm,
    RiscoAvaliacaoForm,
)
from .models import (
    Avaliacao,
    AvaliacaoStatus,
    CategoriaQuestao,
    Empresa,
    LogAuditoriaResposta,
    MetaIndicador,
    ObjetivoEstrategicoPDTI,
    PDTI,
    PlanoAcao,
    PlanoAcaoStatus,
    Questao,
    Resposta,
    RespostaEscolha,
    RiscoAvaliacao,
    UserRole,
)
from .services import (
    gerar_relatorio,
    progresso_avaliacao,
    registrar_log_resposta,
)


def _usuario_acessa_avaliacao(usuario, avaliacao):
    perfil = getattr(usuario, "profile", None)
    if not perfil:
        return False

    if perfil.role == UserRole.ADMIN:
        return True

    if perfil.role == UserRole.CONSULTOR:
        return (
            avaliacao.consultor_responsavel_id == usuario.id
            or avaliacao.participantes.filter(id=usuario.id).exists()
        )

    return avaliacao.participantes.filter(id=usuario.id).exists()


def _usuario_gerencia_avaliacao(usuario, avaliacao):
    perfil = getattr(usuario, "profile", None)
    if not perfil:
        return False

    if perfil.role == UserRole.ADMIN:
        return True

    return perfil.role == UserRole.CONSULTOR and avaliacao.consultor_responsavel_id == usuario.id


@login_required
def dashboard(request):
    perfil = getattr(request.user, "profile", None)
    if not perfil:
        messages.error(request, "Seu usuário não possui perfil definido.")
        return redirect("logout")

    avaliacoes = Avaliacao.objects.select_related("empresa", "consultor_responsavel")
    if perfil.role != UserRole.ADMIN:
        avaliacoes = avaliacoes.filter(participantes=request.user) | avaliacoes.filter(
            consultor_responsavel=request.user
        )

    context = {
        "perfil": perfil,
        "empresas_total": Empresa.objects.count(),
        "questoes_total": Questao.objects.filter(ativa=True).count(),
        "avaliacoes": avaliacoes.distinct()[:10],
    }
    return render(request, "avaliacao/dashboard.html", context)


@role_required(UserRole.ADMIN, UserRole.CONSULTOR)
def empresa_list(request):
    empresas = Empresa.objects.all().order_by("nome")
    return render(request, "avaliacao/empresa_list.html", {"empresas": empresas})


@role_required(UserRole.ADMIN, UserRole.CONSULTOR)
@require_http_methods(["GET", "POST"])
def empresa_create(request):
    if request.method == "POST":
        form = EmpresaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Empresa cadastrada.")
            return redirect("empresa_list")
    else:
        form = EmpresaForm()
    return render(request, "avaliacao/form.html", {"form": form, "titulo": "Nova Empresa"})


@role_required(UserRole.ADMIN)
def questao_list(request):
    questoes = Questao.objects.select_related("categoria").all().order_by("categoria__nome", "id")
    return render(request, "avaliacao/questao_list.html", {"questoes": questoes})


@role_required(UserRole.ADMIN)
def categoria_list(request):
    categorias = CategoriaQuestao.objects.all().order_by("nome")
    return render(request, "avaliacao/categoria_list.html", {"categorias": categorias})


@role_required(UserRole.ADMIN)
@require_http_methods(["GET", "POST"])
def categoria_create(request):
    if request.method == "POST":
        form = CategoriaQuestaoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Categoria cadastrada.")
            return redirect("categoria_list")
    else:
        form = CategoriaQuestaoForm()
    return render(request, "avaliacao/form.html", {"form": form, "titulo": "Nova Categoria"})


@role_required(UserRole.ADMIN)
@require_http_methods(["GET", "POST"])
def questao_create(request):
    if request.method == "POST":
        form = QuestaoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Questão cadastrada.")
            return redirect("questao_list")
    else:
        form = QuestaoForm()
    return render(request, "avaliacao/form.html", {"form": form, "titulo": "Nova Questão"})


@role_required(UserRole.ADMIN)
@require_http_methods(["GET", "POST"])
def questao_update(request, questao_id):
    questao = get_object_or_404(Questao, id=questao_id)
    if request.method == "POST":
        form = QuestaoForm(request.POST, instance=questao)
        if form.is_valid():
            form.save()
            messages.success(request, "Questão atualizada.")
            return redirect("questao_list")
    else:
        form = QuestaoForm(instance=questao)
    return render(request, "avaliacao/form.html", {"form": form, "titulo": "Editar Questão"})


@login_required
def avaliacao_list(request):
    perfil = getattr(request.user, "profile", None)
    if not perfil:
        messages.error(request, "Seu usuário não possui perfil definido.")
        return redirect("logout")

    avaliacoes = Avaliacao.objects.select_related("empresa", "consultor_responsavel")

    if perfil.role == UserRole.ADMIN:
        pass
    elif perfil.role == UserRole.CONSULTOR:
        avaliacoes = avaliacoes.filter(
            Q(consultor_responsavel=request.user) | Q(participantes=request.user)
        ).distinct()
    else:
        avaliacoes = avaliacoes.filter(participantes=request.user).distinct()

    return render(request, "avaliacao/avaliacao_list.html", {"avaliacoes": avaliacoes})


@role_required(UserRole.ADMIN, UserRole.CONSULTOR)
@require_http_methods(["GET", "POST"])
def avaliacao_create(request):
    if request.method == "POST":
        form = AvaliacaoForm(request.POST)
        if form.is_valid():
            avaliacao = form.save()
            avaliacao.participantes.add(avaliacao.consultor_responsavel)
            messages.success(request, "Avaliação criada.")
            return redirect("avaliacao_detail", avaliacao_id=avaliacao.id)
    else:
        form = AvaliacaoForm(initial={"consultor_responsavel": request.user})
    return render(request, "avaliacao/form.html", {"form": form, "titulo": "Nova Avaliação"})


@login_required
def avaliacao_detail(request, avaliacao_id):
    avaliacao = get_object_or_404(Avaliacao, id=avaliacao_id)
    if not _usuario_acessa_avaliacao(request.user, avaliacao):
        messages.error(request, "Você não tem acesso a esta avaliação.")
        return redirect("dashboard")

    questoes = Questao.objects.select_related("categoria").filter(ativa=True).order_by("categoria__nome", "id")
    respostas = {
        resposta.questao_id: resposta
        for resposta in Resposta.objects.filter(avaliacao=avaliacao).select_related("respondido_por", "plano_acao")
    }

    return render(
        request,
        "avaliacao/avaliacao_detail.html",
        {
            "avaliacao": avaliacao,
            "questoes": questoes,
            "respostas": respostas,
            "progresso": progresso_avaliacao(avaliacao),
        },
    )


@login_required
@require_http_methods(["GET", "POST"])
def responder_questao(request, avaliacao_id, questao_id):
    avaliacao = get_object_or_404(Avaliacao, id=avaliacao_id)
    questao = get_object_or_404(Questao, id=questao_id, ativa=True)

    if not _usuario_acessa_avaliacao(request.user, avaliacao):
        messages.error(request, "Você não tem acesso para responder esta avaliação.")
        return redirect("dashboard")

    if avaliacao.status == AvaliacaoStatus.CONCLUIDA:
        messages.error(request, "Avaliações concluídas não podem mais receber alterações.")
        return redirect("avaliacao_detail", avaliacao_id=avaliacao.id)

    resposta = Resposta.objects.filter(avaliacao=avaliacao, questao=questao).first()

    if request.method == "POST":
        form = RespostaForm(request.POST, request.FILES, instance=resposta)
        if form.is_valid():
            registro = form.save(commit=False)
            registro.avaliacao = avaliacao
            registro.questao = questao
            registro.respondido_por = request.user
            registro.save()
            registrar_log_resposta(registro, request.user)

            if registro.resposta == RespostaEscolha.NAO and registro.providencia:
                PlanoAcao.objects.get_or_create(resposta=registro)
            else:
                PlanoAcao.objects.filter(resposta=registro).delete()

            messages.success(request, "Resposta registrada com sucesso.")
            return redirect("avaliacao_detail", avaliacao_id=avaliacao.id)
    else:
        form = RespostaForm(instance=resposta)

    return render(
        request,
        "avaliacao/responder_questao.html",
        {"form": form, "avaliacao": avaliacao, "questao": questao, "resposta": resposta},
    )


@login_required
@require_http_methods(["GET", "POST"])
def plano_acao_update(request, avaliacao_id, resposta_id):
    avaliacao = get_object_or_404(Avaliacao, id=avaliacao_id)
    resposta = get_object_or_404(Resposta, id=resposta_id, avaliacao=avaliacao, resposta=RespostaEscolha.NAO)

    if not _usuario_gerencia_avaliacao(request.user, avaliacao):
        messages.error(request, "Você não tem permissão para editar o plano de ação.")
        return redirect("dashboard")

    plano, _ = PlanoAcao.objects.get_or_create(resposta=resposta)

    if request.method == "POST":
        form = PlanoAcaoForm(request.POST, instance=plano)
        if form.is_valid():
            form.save()
            messages.success(request, "Plano de ação atualizado.")
            return redirect("avaliacao_detail", avaliacao_id=avaliacao.id)
    else:
        form = PlanoAcaoForm(instance=plano)

    return render(
        request,
        "avaliacao/form.html",
        {
            "form": form,
            "titulo": f"Plano de ação - Questão #{resposta.questao.id}",
        },
    )


@login_required
def relatorio(request, avaliacao_id):
    avaliacao = get_object_or_404(Avaliacao, id=avaliacao_id)
    if not _usuario_acessa_avaliacao(request.user, avaliacao):
        messages.error(request, "Você não tem acesso a este relatório.")
        return redirect("dashboard")

    dados = gerar_relatorio(avaliacao)
    return render(request, "avaliacao/relatorio.html", {"avaliacao": avaliacao, **dados})


@role_required(UserRole.ADMIN, UserRole.CONSULTOR)
@require_POST
def concluir_avaliacao(request, avaliacao_id):
    avaliacao = get_object_or_404(Avaliacao, id=avaliacao_id)
    if not _usuario_gerencia_avaliacao(request.user, avaliacao):
        messages.error(request, "Você não tem permissão para concluir esta avaliação.")
        return redirect("dashboard")

    if avaliacao.status == AvaliacaoStatus.CONCLUIDA:
        messages.info(request, "A avaliação já estava concluída.")
    else:
        avaliacao.status = AvaliacaoStatus.CONCLUIDA
        avaliacao.save(update_fields=["status"])
        messages.success(request, "Avaliação marcada como concluída.")
    return redirect("avaliacao_detail", avaliacao_id=avaliacao.id)


@role_required(UserRole.ADMIN, UserRole.CONSULTOR)
def auditoria(request, avaliacao_id):
    avaliacao = get_object_or_404(Avaliacao, id=avaliacao_id)
    if not _usuario_gerencia_avaliacao(request.user, avaliacao):
        messages.error(request, "Você não tem permissão para consultar a auditoria desta avaliação.")
        return redirect("dashboard")

    logs = LogAuditoriaResposta.objects.filter(resposta_registro__avaliacao=avaliacao).select_related(
        "usuario", "resposta_registro__questao"
    )
    return render(request, "avaliacao/auditoria.html", {"avaliacao": avaliacao, "logs": logs})


@role_required(UserRole.ADMIN, UserRole.CONSULTOR)
@require_http_methods(["GET", "POST"])
def risco_create(request, avaliacao_id):
    avaliacao = get_object_or_404(Avaliacao, id=avaliacao_id)

    if not _usuario_gerencia_avaliacao(request.user, avaliacao):
        messages.error(request, "Você não tem permissão para registrar riscos nesta avaliação.")
        return redirect("dashboard")

    if request.method == "POST":
        form = RiscoAvaliacaoForm(request.POST)
        if form.is_valid():
            risco = form.save(commit=False)
            risco.avaliacao = avaliacao
            risco.save()
            messages.success(request, "Risco registrado com sucesso.")
            return redirect("matriz_risco", avaliacao_id=avaliacao.id)
    else:
        form = RiscoAvaliacaoForm()

    return render(request, "avaliacao/form.html", {"form": form, "titulo": "Novo risco"})


@login_required
def matriz_risco(request, avaliacao_id):
    avaliacao = get_object_or_404(Avaliacao, id=avaliacao_id)
    if not _usuario_acessa_avaliacao(request.user, avaliacao):
        messages.error(request, "Você não tem acesso à matriz de riscos desta avaliação.")
        return redirect("dashboard")

    riscos = avaliacao.riscos.select_related("responsavel").all()

    return render(
        request,
        "avaliacao/matriz_risco.html",
        {
            "avaliacao": avaliacao,
            "riscos": riscos,
        },
    )


@login_required
def api_relatorio(request, avaliacao_id):
    avaliacao = get_object_or_404(Avaliacao, id=avaliacao_id)
    if not _usuario_acessa_avaliacao(request.user, avaliacao):
        return JsonResponse({"detail": "forbidden"}, status=403)

    dados = gerar_relatorio(avaliacao)

    payload = {
        "avaliacao": {
            "id": avaliacao.id,
            "nome": avaliacao.nome,
            "empresa": avaliacao.empresa.nome,
            "setor": avaliacao.empresa.setor,
            "status": avaliacao.status,
            "criada_em": avaliacao.criada_em.isoformat(),
        },
        "score_geral": dados["score_geral"],
        "classificacao": dados["classificacao"],
        "score_categoria": dados["score_categoria"],
        "benchmark": dados["benchmark"],
        "historico": [
            {
                "id": item["id"],
                "nome": item["nome"],
                "score": item["score"],
                "status": item["status"],
                "criada_em": item["criada_em"].isoformat(),
            }
            for item in dados["historico"]
        ],
    }
    return JsonResponse(payload)


# ─────────────────────────────────────────────
# PLANO DE AÇÃO 5W2H — página consolidada
# ─────────────────────────────────────────────

@role_required(UserRole.ADMIN, UserRole.CONSULTOR)
@require_http_methods(["GET", "POST"])
def plano_acao_5w2h(request, avaliacao_id):
    avaliacao = get_object_or_404(Avaliacao, id=avaliacao_id)

    if not _usuario_gerencia_avaliacao(request.user, avaliacao):
        messages.error(request, "Você não tem permissão para gerenciar o plano de ação.")
        return redirect("dashboard")

    # Busca todas as respostas NAO que possuem providência (base do plano)
    respostas_nao = (
        Resposta.objects.filter(avaliacao=avaliacao, resposta=RespostaEscolha.NAO)
        .exclude(providencia="")
        .select_related("questao__categoria", "questao")
        .order_by("questao__categoria__nome", "questao__id")
    )

    # Garante que cada resposta tem um PlanoAcao associado
    for r in respostas_nao:
        PlanoAcao.objects.get_or_create(resposta=r)

    if request.method == "POST":
        erros = False
        for r in respostas_nao:
            plano = r.plano_acao
            prefix = f"plano_{plano.pk}"
            form = PlanoAcaoInlineForm(request.POST, instance=plano, prefix=prefix)
            if form.is_valid():
                form.save()
            else:
                erros = True

        if erros:
            messages.warning(request, "Alguns campos não foram salvos. Verifique os erros.")
        else:
            messages.success(request, "Plano de ação atualizado com sucesso!")
        return redirect("plano_acao_5w2h", avaliacao_id=avaliacao_id)

    # Monta lista de (resposta, form) para o template
    itens = []
    for r in respostas_nao:
        plano = r.plano_acao
        form = PlanoAcaoInlineForm(instance=plano, prefix=f"plano_{plano.pk}")
        itens.append({"resposta": r, "plano": plano, "form": form})

    return render(request, "avaliacao/plano_acao_5w2h.html", {
        "avaliacao": avaliacao,
        "itens": itens,
    })


# ─────────────────────────────────────────────
# EXPORTAÇÃO 5W2H → .xlsx
# ─────────────────────────────────────────────

@role_required(UserRole.ADMIN, UserRole.CONSULTOR)
def exportar_plano_xlsx(request, avaliacao_id):
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        messages.error(request, "Biblioteca openpyxl não instalada. Instale com: pip install openpyxl")
        return redirect("plano_acao_5w2h", avaliacao_id=avaliacao_id)

    avaliacao = get_object_or_404(Avaliacao, id=avaliacao_id)

    if not _usuario_acessa_avaliacao(request.user, avaliacao):
        messages.error(request, "Você não tem acesso a esta avaliação.")
        return redirect("dashboard")

    respostas_nao = (
        Resposta.objects.filter(avaliacao=avaliacao, resposta=RespostaEscolha.NAO)
        .exclude(providencia="")
        .select_related("questao__categoria", "questao", "plano_acao", "plano_acao__responsavel")
        .order_by("questao__categoria__nome", "questao__id")
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plano de Ação 5W2H"

    # Estilos
    header_fill = PatternFill("solid", fgColor="1A3A5C")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=13)
    wrap = Alignment(wrap_text=True, vertical="top")

    # Título
    ws.merge_cells("A1:L1")
    ws["A1"] = f"Plano de Ação 5W2H — {avaliacao.empresa.nome} — {avaliacao.nome}"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Cabeçalhos
    headers = [
        "Departamento / Área",
        "What (O que?)",
        "Why (Por quê?)",
        "When (Quando?)",
        "Where (Onde?)",
        "Who (Quem?)",
        "How (Como?)",
        "How Much (Custo/Esforço)",
        "Custo (R$)",
        "Natureza",
        "Recorrência",
        "Status",
        "Framework",
    ]
    ws.append(headers)
    for col, _ in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30

    # Dados
    for r in respostas_nao:
        plano = getattr(r, "plano_acao", None)
        responsavel_nome = plano.responsavel.get_full_name() or plano.responsavel.username if (plano and plano.responsavel) else "-"
        prazo = str(plano.data_limite) if (plano and plano.data_limite) else "-"
        status_display = plano.get_status_display() if plano else "Pendente"
        where_val = plano.where_local if plano else "-"
        how_val = plano.how if plano else "-"
        how_much_val = plano.how_much if plano else "-"
        custo_val = float(plano.custo_valor) if (plano and plano.custo_valor is not None) else "-"
        natureza_val = plano.custo_natureza if plano else "-"
        recorr_val = plano.get_custo_recorrencia_display() if plano else "-"

        row = [
            r.questao.categoria.nome,
            r.questao.texto,
            r.providencia,
            prazo,
            where_val or "-",
            responsavel_nome,
            how_val or "-",
            how_much_val or "-",
            custo_val,
            natureza_val or "-",
            recorr_val or "-",
            status_display,
            r.questao.get_framework_origem_display(),
        ]
        ws.append(row)
        for col in range(1, len(headers) + 1):
            ws.cell(row=ws.max_row, column=col).alignment = wrap

    # Larguras
    col_widths = [22, 45, 45, 14, 25, 22, 40, 22, 14, 12, 14, 16, 16]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # Resposta em buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"plano_acao_5w2h_{avaliacao.empresa.nome.replace(' ', '_')}_{avaliacao_id}.xlsx"
    response = HttpResponse(
        buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ─────────────────────────────────────────────
# EXPORTAÇÃO RELATÓRIO → PDF (WeasyPrint)
# ─────────────────────────────────────────────

@login_required
def exportar_relatorio_pdf(request, avaliacao_id):
    avaliacao = get_object_or_404(Avaliacao, id=avaliacao_id)

    if not _usuario_acessa_avaliacao(request.user, avaliacao):
        messages.error(request, "Você não tem acesso a este relatório.")
        return redirect("dashboard")

    try:
        from weasyprint import HTML, CSS
    except ImportError:
        messages.error(request, "WeasyPrint não está instalado corretamente.")
        return redirect("relatorio", avaliacao_id=avaliacao_id)

    dados = gerar_relatorio(avaliacao)

    html_string = render_to_string(
        "avaliacao/relatorio_pdf.html",
        {"avaliacao": avaliacao, "request": request, **dados},
    )

    pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri("/")).write_pdf()

    filename = f"relatorio_sam_ti_{avaliacao.empresa.nome.replace(' ', '_')}_{avaliacao_id}.pdf"
    response = HttpResponse(pdf_file, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ─────────────────────────────────────────────
# METAS DE FUTURO (2029) — por Empresa
# ─────────────────────────────────────────────

@role_required(UserRole.ADMIN, UserRole.CONSULTOR)
@require_http_methods(["GET"])
def metas_2029_list(request, empresa_id):
    empresa = get_object_or_404(Empresa, id=empresa_id)

    if request.user.profile.role != UserRole.ADMIN and empresa.owner_id != request.user.id:
        # consultor pode acessar via avaliação; aqui mantemos simples e restrito
        pass

    metas = MetaIndicador.objects.filter(empresa=empresa)
    return render(request, "avaliacao/metas_2029_list.html", {"empresa": empresa, "metas": metas})


@role_required(UserRole.ADMIN, UserRole.CONSULTOR)
@require_http_methods(["GET", "POST"])
def metas_2029_create(request, empresa_id):
    empresa = get_object_or_404(Empresa, id=empresa_id)

    if request.method == "POST":
        form = MetaIndicadorForm(request.POST)
        if form.is_valid():
            meta = form.save(commit=False)
            meta.empresa = empresa
            meta.save()
            messages.success(request, "Meta criada.")
            return redirect("metas_2029_list", empresa_id=empresa.id)
    else:
        form = MetaIndicadorForm()

    return render(request, "avaliacao/form.html", {"form": form, "titulo": "Nova meta (2029)"})


@role_required(UserRole.ADMIN, UserRole.CONSULTOR)
@require_http_methods(["GET", "POST"])
def metas_2029_update(request, empresa_id, meta_id):
    empresa = get_object_or_404(Empresa, id=empresa_id)
    meta = get_object_or_404(MetaIndicador, id=meta_id, empresa=empresa)

    if request.method == "POST":
        form = MetaIndicadorForm(request.POST, instance=meta)
        if form.is_valid():
            form.save()
            messages.success(request, "Meta atualizada.")
            return redirect("metas_2029_list", empresa_id=empresa.id)
    else:
        form = MetaIndicadorForm(instance=meta)

    return render(request, "avaliacao/form.html", {"form": form, "titulo": "Editar meta (2029)"})


@role_required(UserRole.ADMIN, UserRole.CONSULTOR)
@require_http_methods(["GET", "POST"])
def metas_2029_delete(request, empresa_id, meta_id):
    empresa = get_object_or_404(Empresa, id=empresa_id)
    meta = get_object_or_404(MetaIndicador, id=meta_id, empresa=empresa)

    if request.method == "POST":
        meta.delete()
        messages.success(request, "Meta removida.")
        return redirect("metas_2029_list", empresa_id=empresa.id)

    return render(
        request,
        "avaliacao/confirm_delete.html",
        {"titulo": "Excluir meta", "descricao": meta.nome_indicador, "voltar_url": redirect("metas_2029_list", empresa_id=empresa.id).url},
    )


# ─────────────────────────────────────────────
# PDTI — criar/editar/visualizar por Avaliação (somente quando CONCLUÍDA)
# ─────────────────────────────────────────────

@login_required
@role_required(UserRole.ADMIN, UserRole.CONSULTOR)
@require_http_methods(["GET", "POST"])
def pdti_view(request, avaliacao_id):
    avaliacao = get_object_or_404(Avaliacao, id=avaliacao_id)

    if not _usuario_acessa_avaliacao(request.user, avaliacao):
        messages.error(request, "Você não tem acesso.")
        return redirect("dashboard")

    pdti, _ = PDTI.objects.get_or_create(avaliacao=avaliacao)

    pode_editar = _usuario_gerencia_avaliacao(request.user, avaliacao) and avaliacao.status == AvaliacaoStatus.CONCLUIDA

    if request.method == "POST":
        if not pode_editar:
            messages.error(request, "O PDTI só pode ser editado após a avaliação estar concluída.")
            return redirect("pdti_view", avaliacao_id=avaliacao.id)

        form = PDTIForm(request.POST, instance=pdti)
        if form.is_valid():
            form.save()
            messages.success(request, "PDTI salvo.")
            return redirect("pdti_view", avaliacao_id=avaliacao.id)
    else:
        form = PDTIForm(instance=pdti)

    objetivos = ObjetivoEstrategicoPDTI.objects.filter(empresa=avaliacao.empresa)
    metas = MetaIndicador.objects.filter(empresa=avaliacao.empresa)

    dados = gerar_relatorio(avaliacao)

    return render(
        request,
        "avaliacao/pdti.html",
        {
            "avaliacao": avaliacao,
            "empresa": avaliacao.empresa,
            "pdti": pdti,
            "form": form,
            "pode_editar": pode_editar,
            "objetivos": objetivos,
            "metas": metas,
            "relatorio": dados,
        },
    )


@login_required
@role_required(UserRole.ADMIN, UserRole.CONSULTOR)
def exportar_pdti_pdf(request, avaliacao_id):
    avaliacao = get_object_or_404(Avaliacao, id=avaliacao_id)

    if not _usuario_acessa_avaliacao(request.user, avaliacao):
        messages.error(request, "Você não tem acesso.")
        return redirect("dashboard")

    try:
        from weasyprint import HTML
    except ImportError:
        messages.error(request, "WeasyPrint não está instalado corretamente.")
        return redirect("pdti_view", avaliacao_id=avaliacao_id)

    pdti, _ = PDTI.objects.get_or_create(avaliacao=avaliacao)
    objetivos = ObjetivoEstrategicoPDTI.objects.filter(empresa=avaliacao.empresa)
    metas = MetaIndicador.objects.filter(empresa=avaliacao.empresa)

    dados = gerar_relatorio(avaliacao)

    html_string = render_to_string(
        "avaliacao/pdti_pdf.html",
        {
            "avaliacao": avaliacao,
            "empresa": avaliacao.empresa,
            "pdti": pdti,
            "objetivos": objetivos,
            "metas": metas,
            "respostas_nao": Resposta.objects.filter(avaliacao=avaliacao, resposta=RespostaEscolha.NAO)
            .exclude(providencia="")
            .select_related("questao__categoria", "plano_acao", "plano_acao__responsavel"),
            **dados,
        },
    )

    pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri("/")).write_pdf()

    filename = f"pdti_{avaliacao.empresa.nome.replace(' ', '_')}_{avaliacao_id}.pdf"
    response = HttpResponse(pdf_file, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
